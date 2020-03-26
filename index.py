import win32api
import win32print
import time
import glob
from openpyxl import load_workbook
from docxtpl import DocxTemplate
import datetime
import os


def printdoc(filename):
    handle = win32print.OpenPrinter(win32print.GetDefaultPrinter())
    status1 = 1024  # 初始化为1024表示忙 0是空闲状态
    status2 = 2  # 打印机队列数
    while (status1 == 1024) and (status2 >= 2):
        statusqueue = (win32print.GetPrinter(handle))
        status1 = statusqueue[-3]
        status2 = statusqueue[-2]
        print("    等待中，打印队列" + str(status2) + "文档")
        time.sleep(1)
        continue
    print("开始打印" + filename)
    win32api.ShellExecute(0, 'print', filename, win32print.GetDefaultPrinterW(), ".", 0)
    while ((win32print.GetPrinter(handle)[-2]) != (status2 +1)):
        time.sleep(1)
        print("等待打印机反馈")
    os.remove(filename)


#
# for i in range(1, 10):
#     printdoc(str(i) + ".docx")

test = glob.glob("*.xlsx")
if (test != []):
    print("找到符合要求文件:" + test[0])
    file = test[0]
    wb = load_workbook(file)
    ws = wb.active
    maxrow = ws.max_row
    maxcol = ws.max_column
    cur = datetime.datetime.now()
    date = str(cur.year) + '年' + str(cur.month) + '月' + str(cur.day) + '日'
    print(win32print.GetDefaultPrinter())
    for row in range(2, maxrow + 1):
        tpl = DocxTemplate("毒驾注销告知通知书.docx")
        valuexm = ws.cell(row=row, column=2).value
        valuesfzmhm = ws.cell(row=row, column=4).value
        # valuezjcx = ws.cell(row=row, column=9).value
        change = {'AAA': valuexm, 'BBB': valuesfzmhm, 'DDD': date}  # 在这里修改日期
        tpl.render(change)
        tpl.save("自动生成" + valuesfzmhm + ".docx")
        fn = "自动生成" + valuesfzmhm + ".docx"
        print("生成" + fn)
        printdoc(fn)
    print("队列完成请等待打印完成！")
    input()
else:
    print("请在本文件夹放入xlsx数据文件!(老版xls文件不接受）")
    input()
# print(test)
